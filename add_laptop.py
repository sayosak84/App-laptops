# import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *
import tkinter as tk
from tkinter import ttk
import pandas as pd
import sqlite3
from sqlite3 import Error
import os

# opening the existing excel file
wb = load_workbook('C:\\Users\\Nader\\Documents\\projets_lotfi\\App-laptops\\excel.xlsx')

# create the sheet object
sheet = wb.active

LARGEFONT =("Verdana", 35)
MIDFONT =("Verdana", 22)
to_search = None




def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

conn = sqlite3.connect(resource_path('test.db'))

class tkinterApp(tk.Tk):
     
    # __init__ function for class tkinterApp
    def __init__(self, *args, **kwargs):
         
        # __init__ function for class Tk
        tk.Tk.__init__(self, *args, **kwargs)

        self.title("Gestion des laptops")
        self.geometry("700x500")
        # creating a container
        container = tk.Frame(self) 
        container.pack(side = "top", fill = "both", expand = True)

        container.grid_rowconfigure(0, weight = 3)
        container.grid_columnconfigure(0, weight = 3)

  
        # initializing frames to an empty array
        self.frames = {} 
  
        # iterating through a tuple consisting
        # of the different page layouts
        for F in (StartPage, Page1, Page2):
  
            frame = F(container, self)
  
            # initializing frame of that object from
            # startpage, page1, page2 respectively with
            # for loop
            self.frames[F] = frame
  
            frame.grid(row = 0, column = 0, sticky ="nsew")
  
        self.show_frame(Page1)
  
    # to display the current frame passed as
    # parameter
    def show_frame(self, cont):
        frame = self.frames[cont]
        frame.tkraise()








class StartPage(tk.Frame):
    def __init__(self, parent, controller):

        def excel():
    
            # resize the width of columns in
            # excel spreadsheet
            sheet.column_dimensions['A'].width = 30
            sheet.column_dimensions['B'].width = 10
            sheet.column_dimensions['C'].width = 10
            sheet.column_dimensions['D'].width = 20
            sheet.column_dimensions['E'].width = 20
            sheet.column_dimensions['F'].width = 40
            sheet.column_dimensions['G'].width = 50
            sheet.column_dimensions['H'].width = 50
            sheet.column_dimensions['I'].width = 50
            sheet.column_dimensions['J'].width = 50
            sheet.column_dimensions['K'].width = 50

            # write given data to an excel spreadsheet
            # at particular location
            sheet.cell(row=1, column=1).value = "Nom de l'appareil"
            sheet.cell(row=1, column=2).value = "Marque"
            sheet.cell(row=1, column=3).value = "Date de l'achat"
            sheet.cell(row=1, column=4).value = "Form Number"
            sheet.cell(row=1, column=5).value = "Contact Number"
            sheet.cell(row=1, column=6).value = "Email id"
            sheet.cell(row=1, column=7).value = "Amortissement mensuel"
            sheet.cell(row=1, column=8).value = "Dotation aux amortissements"
            sheet.cell(row=1, column=9).value = "Valeure compatable nette"
            sheet.cell(row=1, column=10).value = "Utilisateur"
            sheet.cell(row=1, column=11).value = "Date affectation"


        # Function to set focus (cursor)
        def focus1(event):
            # set focus on the marque_field box
            marque_field.focus_set()


        # Function to set focus
        def focus2(event):
            # set focus on the date_achat_field box
            date_achat_field.focus_set()


        # Function to set focus
        def focus3(event):
            # set focus on the date_exp_field box
            date_exp_field.focus_set()


        # Function to set focus
        def focus4(event):
            # set focus on the prix_achat_field box
            prix_achat_field.focus_set()


        # Function to set focus
        def focus5(event):
            # set focus on the date_compta_field box
            date_compta_field.focus_set()


        # Function to set focus
        def focus6(event):
            # set focus on the ammort_mensuel_field box
            ammort_mensuel_field.focus_set()

        # Function to set focus
        def focus7(event):
            # set focus on the utilisateur_field box
            utilisateur_field.focus_set()

        # Function to set focus
        def focus8(event):
            # set focus on the date_affect_field box
            date_affect_field.focus_set()


        # Function for clearing the
        # contents of text entry boxes
        def clear():
            
            # clear the content of text entry box
            nom_appareil_field.delete(0, END)
            marque_field.delete(0, END)
            date_achat_field.delete(0, END)
            date_exp_field.delete(0, END)
            prix_achat_field.delete(0, END)
            date_compta_field.delete(0, END)
            ammort_mensuel_field.delete(0, END)
            utilisateur_field.delete(0, END)
            date_affect_field.delete(0, END)


        # Function to take data from GUI
        # window and write to an excel file
        def insert():
            elements = []
            
            # if user not fill any entry
            # then print "empty input"
            if (nom_appareil_field.get() == "" and
                marque_field.get() == "" and
                date_achat_field.get() == "" and
                date_exp_field.get() == "" and
                prix_achat_field.get() == "" and
                date_compta_field.get() == "" and
                ammort_mensuel_field.get() == "" and
                utilisateur_field.get() == "" and
                date_affect_field.get() == ""):
                    
                print("empty input")

            else:

                # assigning the max row and max column
                # value upto which data is written
                # in an excel sheet to the variable
                current_row = sheet.max_row
                current_column = sheet.max_column

                # get method returns current text
                # as string which we write into
                # excel spreadsheet at particular location
                sheet.cell(row=current_row + 1, column=1).value = nom_appareil_field.get()
                sheet.cell(row=current_row + 1, column=2).value = marque_field.get()
                sheet.cell(row=current_row + 1, column=3).value = date_achat_field.get()
                sheet.cell(row=current_row + 1, column=4).value = date_exp_field.get()
                sheet.cell(row=current_row + 1, column=5).value = prix_achat_field.get()
                sheet.cell(row=current_row + 1, column=6).value = date_compta_field.get()
                sheet.cell(row=current_row + 1, column=7).value = ammort_mensuel_field.get()
                sheet.cell(row=current_row + 1, column=10).value = utilisateur_field.get()
                sheet.cell(row=current_row + 1, column=11).value = date_affect_field.get()


                elements.append(nom_appareil_field.get())
                elements.append(marque_field.get())
                elements.append(date_achat_field.get())
                elements.append(date_exp_field.get())
                elements.append(prix_achat_field.get())
                elements.append(date_compta_field.get())
                elements.append(ammort_mensuel_field.get())
                elements.append("test")
                elements.append("test")
                elements.append(utilisateur_field.get())
                elements.append(date_affect_field.get())
                elements.append("test")

                rqst = """INSERT INTO laptops
                          (Nom_appareil, Marque, Date_achat, Form_Number, Contact_Number, Email_id, Amortissement_mensuel, Dotation_amortissements, Valeure_compatable_nette, Utilisateur, Date_affectation, Date_fin) 
                          VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);"""

                cursor = conn.cursor()
                cursor.execute(rqst, elements)
                conn.commit()
                # save the file
                wb.save('C:\\Users\\Nader\\Documents\\projets_lotfi\\App-laptops\\excel.xlsx')

                # set focus on the nom_appareil_field box
                nom_appareil_field.focus_set()

                # call the clear() function
                clear()

        tk.Frame.__init__(self, parent)

        #root = Tk()

            # set the background colour of GUI window
        self.configure(background='light green')

        excel()

            # create a Form label
        heading = Label(self, text="Form", bg="light green")

            # create a nom_appareil label
        nom_appareil = Label(self, text="Nom de l'appareil", bg="light green")

            # create a marque label
        marque = Label(self, text="Marque", bg="light green")

            # create a Date de l'achat label
        date_achat = Label(self, text="Date de l'achat", bg="light green")

            # create a date_exp label
        date_exp = Label(self, text="Date d'expiration", bg="light green")

            # create a prix_achat label
        prix_achat = Label(self, text="Prix d'achat", bg="light green")

            # create a Email id label
        date_compta = Label(self, text="Date comptabilisation", bg="light green")

            # create a ammort_mensuel label
        ammort_mensuel = Label(self, text="Ammortissement mensuel", bg="light green")

            # create a utilisateur label
        utilisateur = Label(self, text="Utilisateur", bg="light green")

            # create a date affectation label
        date_affect = Label(self, text="Date affectation", bg="light green")

            # grid method is used for placing
            # the widgets at respective positions
            # in table like structure .
        heading.grid(row=0, column=1)
        nom_appareil.grid(row=1, column=0)
        marque.grid(row=2, column=0)
        date_achat.grid(row=3, column=0)
        date_exp.grid(row=4, column=0)
        prix_achat.grid(row=5, column=0)
        date_compta.grid(row=6, column=0)
        ammort_mensuel.grid(row=7, column=0)
        utilisateur.grid(row=8, column=0)
        date_affect.grid(row=9, column=0)

        # create a text entry box
        # for typing the information
        nom_appareil_field = Entry(self)
        marque_field = Entry(self)
        date_achat_field = Entry(self)
        date_exp_field = Entry(self)
        prix_achat_field = Entry(self)
        date_compta_field = Entry(self)
        ammort_mensuel_field = Entry(self)
        utilisateur_field = Entry(self)
        date_affect_field = Entry(self)

        # bind method of widget is used for
        # the binding the function with the events

        # whenever the enter key is pressed
        # then call the focus1 function
        nom_appareil_field.bind("<Return>", focus1)

        # whenever the enter key is pressed
        # then call the focus2 function
        marque_field.bind("<Return>", focus2)

        # whenever the enter key is pressed
        # then call the focus3 function
        date_achat_field.bind("<Return>", focus3)

        # whenever the enter key is pressed
        # then call the focus4 function
        date_exp_field.bind("<Return>", focus4)

        # whenever the enter key is pressed
        # then call the focus5 function
        prix_achat_field.bind("<Return>", focus5)

        # whenever the enter key is pressed
        # then call the focus6 function
        date_compta_field.bind("<Return>", focus6)

        ammort_mensuel_field.bind("<Return>", focus7)

        utilisateur_field.bind("<Return>", focus8)



        # grid method is used for placing
        # the widgets at respective positions
        # in table like structure .
        nom_appareil_field.grid(row=1, column=1, ipadx="100")
        marque_field.grid(row=2, column=1, ipadx="100")
        date_achat_field.grid(row=3, column=1, ipadx="100")
        date_exp_field.grid(row=4, column=1, ipadx="100")
        prix_achat_field.grid(row=5, column=1, ipadx="100")
        date_compta_field.grid(row=6, column=1, ipadx="100")
        ammort_mensuel_field.grid(row=7, column=1, ipadx="100")
        utilisateur_field.grid(row=8, column=1, ipadx="100")
        date_affect_field.grid(row=9, column=1, ipadx="100")

        # call excel function
        excel()

        # create a Submit Button and place into the self window
        submit = Button(self, text="Submit", fg="Black",
                                bg="Red", command=insert)
        submit.grid(row=10, column=1)


        button1 = ttk.Button(self, text ="Retour",
                                command = lambda : controller.show_frame(Page1))
     
        # putting the button in its place
        # by using grid
        button1.grid(row = 11, column = 1, padx = 10, pady = 10)



# second window frame page1
class Page1(tk.Frame):
     
    def __init__(self, parent, controller):
         
        tk.Frame.__init__(self, parent)

        logo = tk.PhotoImage(file='C:\\Users\\Nader\\Documents\\projets_lotfi\\App-laptops\\bg-section-1.png')
        BGlabel = tk.Label(self,image=logo)
        BGlabel.image = logo
        BGlabel.place(x=0,y=0, anchor = 'nw')

        logo1 = tk.PhotoImage(file='C:\\Users\\Nader\\Documents\\projets_lotfi\\App-laptops\\logo-DDM.png')
        BGlabel1 = tk.Label(self,image=logo1)
        BGlabel1.image = logo1
        BGlabel1.place(x=0,y=0, anchor = 'nw')

        label = ttk.Label(self, text ="Gestion des laptops", font = LARGEFONT)
        label.grid(row = 0, column = 1, padx = 10, pady = 10)

        
        #label.place(relx=0.5, rely=0.1 , anchor=CENTER)

        #bg_canvas = Canvas(self, width = 700, height=500)
        #bg = PhotoImage(file='C:\\Users\\Nader\\Documents\\projets_lotfi\\App-laptops\\bg-section-1.png')
        #bg_canvas.pack(fill="both", expand=True)
        #bg_canvas.create_image(0,0, image = bg)

        # button to show frame 2 with text
        # layout2
        
        button1 = ttk.Button(self, text ="Ajouter",
                            command = lambda : controller.show_frame(StartPage))
     
        # putting the button in its place
        # by using grid
        button1.grid(row = 1, column = 1, padx = 10, pady = 10)

        # button to show frame 2 with text
        # layout2
        button2 = ttk.Button(self, text ="Rechercher",
                            command = lambda : controller.show_frame(Page2))
     
        # putting the button in its place by
        # using grid
        button2.grid(row = 2, column = 1, padx = 10, pady = 10)

# second window frame page1
class Page2(tk.Frame):
     
    def __init__(self, parent, controller):

        def rechercher():
            df = pd.read_excel('C:\\Users\\Nader\\Documents\\projets_lotfi\\App-laptops\\excel.xlsx')
            result = df[df[df.columns[0]] == "DDM-PF2DXHZS"] # this will only contain 2,4,6 rows
            for row in result.iterrows():
                print("----------------------------------")
                print(row)
                canvas.create_text(400, 100, text = str(row[1]), fill="black", font=('Helvetica 8 bold'))

        tk.Frame.__init__(self, parent)
        label = ttk.Label(self, text ="Recherche", font = MIDFONT)
        label.grid(row = 0, column = 1, padx = 10, pady = 10)
        #label.place(relx=0.5, rely=0.1 , anchor=CENTER)
  
        List_laptops = Label(self, text="liste des laptops", bg="light green")
        List_laptops.grid(row=1, column=0)
        List_laptops_field = Entry(self)
        List_laptops_field.grid(row=1, column=1, ipadx="100")


        canvas= Canvas(self, width= 500, height= 500 , bg="white")
        canvas.grid(row=2, column=1, ipadx="100")

        # button to show frame 2 with text
        # layout2
        button1 = ttk.Button(self, text ="Rechercher", command = rechercher)
     
        # putting the button in its place
        # by using grid
        button1.grid(row = 3, column = 1, padx = 10, pady = 10)
  
        # button to show frame 2 with text
        # layout2
        button2 = ttk.Button(self, text ="Retour",
                            command = lambda : controller.show_frame(Page1))
     
        # putting the button in its place by
        # using grid
        button2.grid(row = 4, column = 1, padx = 10, pady = 10)


    
app = tkinterApp()
app.mainloop()


